from django.core.management.base import BaseCommand, CommandError
from humanresources.models       import ContractProposal, Contract

from openpyxl import Workbook, load_workbook

class Command(BaseCommand):

    def add_arguments(self, parser):
        parser.add_argument('filename', type=str)

    def handle(self, *args, **options):

        filename = options['filename']

        wb = load_workbook(filename)
        ws = wb[wb.sheetnames[0]]

        rows = [row for row in ws.iter_rows()]
        
        last_row = 0
        for i, row in enumerate(rows):
            if row[0].value and row[0].value.startswith("First Name"): 
                last_row = i
                break

        person = object()

        for i, row in enumerate(rows[last_row+1:]):

            first               = row[0].value
            last                = row[1].value

            print(i, first, last)

            position            = row[2].value
            project             = row[3].value
            type       = row[4].value
            ref        = row[5].value
            start      = row[6].value
            end        = row[7].value
            requisition_start   = row[8].value
            requisition_end     = row[9].value
            funding             = row[10].value
            proj                = row[11].value
            salary              = row[12].value
            salary_from_lab     = row[13].value
